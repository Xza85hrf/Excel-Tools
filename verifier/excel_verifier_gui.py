from excel_verifier import main as run_check_files, adjust_column_width
import tkinter as tk
from tkinter import ttk, filedialog, Text, Toplevel, Scrollbar, messagebox, Menu
import os
import threading
import queue
import csv
from openpyxl import Workbook, load_workbook
import tkinter.font as tkfont
import json

# --- Polish Translations (Complete) --- 
LANG_PL = {
    "window_title": "Weryfikator Plików Excela",
    "tooltip_class_background": "#ffffe0",
    "select_file_button": "Przeglądaj...",
    "select_dir_button": "Przeglądaj...",
    "excel_file_types": "Pliki Excel",
    "csv_file_types": "Pliki CSV",
    "all_files": "Wszystkie pliki",
    "error": "Błąd",
    "output_must_be_xlsx": "Plik wynikowy musi być w formacie .xlsx",
    "failed_to_save_file": "Nie udało się zapisać pliku: {e}",
    "ask_select_all": "Proszę wybrać plik Excel, katalog i kolumnę.",
    "ask_excel_exists": "Wybrany plik Excel nie istnieje.",
    "ask_dir_exists": "Wybrany katalog nie istnieje.",
    "save_cancelled": "Zapis anulowany.",
    "starting_check": "Rozpoczynanie sprawdzania...",
    "saving_results": "Zapisywanie wyników...",
    "check_complete_saved": "Sprawdzanie zakończone! Wyniki zapisano do {filename}",
    "check_complete_save_failed": "Sprawdzanie zakończone, ale nie udało się zapisać pliku.",
    "error_during_process": "Błąd podczas przetwarzania: {e}",
    "log_window_title": "excel_comparison.log",
    "log_view_button": "Pokaż Log", 
    "could_not_open_log": "Nie można otworzyć pliku excel_comparison.log",
    "failed_to_load_columns": "Nie udało się wczytać kolumn: {e}",
    "help_button": "Pomoc",
    "show_help_button": "Pokaż Pomoc",
    "help_title": "Pomoc",
    "help_text": (
        "**Krok po Kroku:**\n\n" # Heading
        "1.  **Wybierz Plik Excel:** Wybierz plik `.xlsx` zawierający listę elementów (np. faktur) do sprawdzenia.\n" # Step 1
        "2.  **Wybierz Katalog:** Wybierz folder, w którym znajdują się pliki do porównania.\n" # Step 2
        "3.  **Wprowadź Kolumnę:** Podaj literę kolumny (np. 'A') z nazwami/ID w Excelu.\n" # Step 3
        "4.  **Wprowadź Rozszerzenie:** Podaj typ plików (np. `.pdf`, `.txt`) do wyszukania.\n" # Step 4
        "5.  **Wybierz Strategię:** Wybierz metodę porównywania nazw (szczegóły poniżej).\n" # Step 5
        "6.  **Uwzględnij Podkatalogi:** Zaznacz, aby szukać również w podfolderach.\n" # Step 6
        "7.  **Kliknij 'Sprawdź Pliki':** Rozpocznij proces sprawdzania.\n\n" # Step 7
        "**Strategie Dopasowania:**\n\n" # Heading
        "*   **Dokładna:** Nazwa pliku musi dokładnie odpowiadać wpisowi w Excelu (bez rozszerzenia, wielkość liter ignorowana).\n" # Strategy 1
        "*   **Zawiera:** Nazwa pliku musi zawierać wpis z Excela (wielkość liter ignorowana).\n\n" # Strategy 2
        "**Wyniki i Funkcje:**\n\n" # Heading
        "*   **Tabela Wyników:** Pokazuje status 'Znaleziony' lub 'Brakujący' dla każdego elementu.\n" # Feature 1
        "*   **Sugestie (Podwójne Kliknięcie):** Dla 'Brakujących', podwójne kliknięcie może pokazać podobne pliki.\n" # Feature 2
        "*   **Eksport CSV:** Zapisz wyniki widoczne w tabeli do pliku CSV.\n" # Feature 3
        "*   **Logi:** Sprawdź szczegóły operacji lub błędy (Plik -> Zobacz Log).\n" # Feature 4
    ),
    "export_csv_button": "Eksportuj Tabelę Wyników jako CSV",
    "export_csv_no_results": "Brak wyników do wyeksportowania.",
    "export_csv_success": "Eksport CSV zakończony: {filename}",
    "export_csv_failed": "Nie udało się zapisać CSV: {e}",
    "status_ready": "Gotowy",
    "status_processing_complete": "Przetwarzanie zakończone. Sprawdzono {count} elementów.",
    "suggestions_title": "Sugestie",
    "suggestions_found": "Możliwe dopasowania dla '{value}':\n- {suggestions}",
    "suggestions_not_found": "Nie znaleziono konkretnych sugestii dla '{value}'.",
    "file_menu": "Plik",
    "view_log_menu": "Zobacz Log",
    "exit_menu": "Wyjście",
    "inputs_label": "Dane Wejściowe",
    "excel_file_label": "Plik Excel z Listą:",
    "dir_label": "Katalog do Sprawdzenia:",
    "options_label": "Opcje",
    "column_label": "Kolumna z Excela:",
    "ext_label": "Rozszerzenie Pliku:",
    "strategy_label": "Strategia Dopasowania:",
    "include_subdirs_label": "Uwzględnij Podkatalogi",
    "check_files_button": "Sprawdź Pliki i Zapisz Wyniki Jako...",
    "results_label": "Wyniki",
    "results_item_header": "Wartość Elementu",
    "results_status_header": "Status",
    "results_details_header": "Znaleziona Ścieżka / Sugestie",
    "status_found": "Znaleziono",
    "status_missing": "Brakujący",
    "tooltip_excel_entry": "Wprowadź ścieżkę do pliku .xlsx lub .csv, lub użyj przycisku Przeglądaj.",
    "tooltip_excel_button": "Wybierz plik Excel (.xlsx/.csv) z listą kontrolną.",
    "tooltip_dir_entry": "Wprowadź ścieżkę do katalogu zawierającego pliki do sprawdzenia, lub użyj przycisku Przeglądaj.",
    "tooltip_dir_button": "Wybierz katalog do przeszukania.",
    "tooltip_column_combo": "Wybierz kolumnę w pliku Excel zawierającą elementy do dopasowania.",
    "tooltip_ext_combo": "Wybierz typ rozszerzenia plików do wyszukania.",
    "tooltip_strategy_combo": "'Dokładna': Nazwa pliku musi pasować do wpisu Excel.\n'Zawiera': Nazwa pliku musi zawierać wpis Excel.",
    "tooltip_subdir_check": "Zaznacz to pole, aby przeszukiwać foldery wewnątrz wybranego katalogu.",
    "tooltip_check_button": "Rozpocznij proces sprawdzania i zapisz wyniki do nowego pliku Excel.",
    "tooltip_export_button": "Zapisz wyniki widoczne w tabeli do pliku CSV.",
    "config_load_error": "Nie udało się wczytać konfiguracji: {e}",
    "config_save_error": "Nie udało się zapisać konfiguracji: {e}",
    "save_suffix": " - Sprawdzone Wyniki",
    "strategy_exact": "Dokładna",
    "strategy_contains": "Zawiera",
    "strategy_startswith": "Zaczyna się od",
    "strategy_regex": "Wyrażenie regularne",
}

# --- Strategy Mapping --- 
STRATEGY_MAP_PL = {
    LANG_PL["strategy_exact"]: 'exact',
    LANG_PL["strategy_contains"]: 'contains',
}

# --- Utility function to get key from value --- 
def get_key_from_value(d, val):
    for key, v in d.items():
        if v == val:
            return key
    return None

class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event=None):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25

        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")

        label = tk.Label(self.tooltip_window, text=self.text, justify='left',
                         background=LANG_PL["tooltip_class_background"], relief='solid', borderwidth=1,
                         font=("Segoe UI", 10, "normal"))
        label.pack(ipadx=5, ipady=3)

    def hide_tooltip(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
        self.tooltip_window = None

def select_file(entry):
    filename = filedialog.askopenfilename(filetypes=[(LANG_PL["excel_file_types"], '*.xlsx *.csv'), (LANG_PL["all_files"], '*.*')])
    if filename:
        entry.delete(0, tk.END)
        entry.insert(0, filename)
        populate_columns(filename)
        update_check_button()

def select_directory(entry):
    directory = filedialog.askdirectory()
    if directory:
        entry.delete(0, tk.END)
        entry.insert(0, directory)
        update_check_button()

def save_to_file(workbook, output_file):
    _, ext = os.path.splitext(output_file)
    if ext.lower() != ".xlsx":
        messagebox.showerror(LANG_PL["error"], LANG_PL["output_must_be_xlsx"])
        return False
    try:
        if workbook.active:
             adjust_column_width(workbook.active)
        workbook.save(output_file)
        return True
    except Exception as e:
        messagebox.showerror(LANG_PL["error"], LANG_PL["failed_to_save_file"].format(e=e))
        return False

def start_file_check(excel_file_entry, directory_entry, q):
    excel_file = excel_file_entry.get()
    directory = directory_entry.get()
    col_name = column_combobox.get()
    extensions = ext_var.get()
    include_subdirs = recurse_var.get()
    match_mode_display = strategy_var.get()
    match_mode_internal = STRATEGY_MAP_PL.get(match_mode_display, 'exact') # Get internal key

    if not excel_file or not directory or not col_name:
        q.put(("update_status", LANG_PL["ask_select_all"]))
        return
    if not os.path.isfile(excel_file):
        q.put(("update_status", LANG_PL["ask_excel_exists"]))
        return
    if not os.path.isdir(directory):
        q.put(("update_status", LANG_PL["ask_dir_exists"]))
        return

    # --- Suggest Output Filename --- 
    base_name, _ = os.path.splitext(os.path.basename(excel_file))
    suggested_filename = f"{base_name}{LANG_PL['save_suffix']}.xlsx"

    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                               filetypes=[(LANG_PL["excel_file_types"], '*.xlsx')],
                                               initialfile=suggested_filename) # Suggest name
    if not output_file:
        q.put(("update_status", LANG_PL["save_cancelled"]))
        return # Exit if no output file selected

    check_files_button['state'] = 'disabled'
    progress['mode'] = 'indeterminate'
    progress.start()
    q.put(("update_status", LANG_PL["starting_check"]))
    threading.Thread(target=run_check_process,
                     args=(excel_file, directory, col_name, extensions, include_subdirs, match_mode_internal, q, output_file)).start()

def run_check_process(excel_file, directory, col_name, extensions, include_subdirs, match_mode, q, output_file):
    try:
        success, workbook = run_check_files(excel_file, directory, col_name, extensions, include_subdirs, match_mode, q)
        if success and workbook is not None:
            q.put(("update_status", LANG_PL["saving_results"])) # Status update handled by excel_comparison now
            if save_to_file(workbook, output_file):
                # Final status update comes from excel_comparison
                pass # q.put(("update_status", f"Check complete! Output saved to {os.path.basename(output_file)}"))
            else:
                # Error shown by save_to_file
                q.put(("update_status", LANG_PL["check_complete_save_failed"]))
    except Exception as e:
        q.put(("update_status", LANG_PL["error_during_process"].format(e=str(e))))
    finally:
        q.put(("update_button", "normal"))
        q.put(("stop_progress", None))

def update_gui(q):
    try:
        while True:
            msg = q.get(0)
            cmd, arg = msg
            if cmd == "update_status": # Renamed from update_label
                status_label['text'] = arg
            elif cmd == "update_progress": # Handle progress updates
                progress['mode'] = 'determinate'
                progress['value'] = arg
            elif cmd == "stop_progress":
                 progress.stop()
                 progress['mode'] = 'determinate' # Reset to determinate mode
                 progress['value'] = 0
            elif cmd == "update_button":
                check_files_button['state'] = arg
            elif cmd == "results":
                show_results(arg)
    except queue.Empty:
        pass
    finally:
        root.after(100, update_gui, q)

def display_log():
    log_window = Toplevel(root)
    log_window.title(LANG_PL["log_window_title"])
    log_window.geometry("600x400")
    log_frame = ttk.Frame(log_window, padding=10)
    log_frame.pack(expand=True, fill='both')
    log_frame.rowconfigure(0, weight=1)
    log_frame.columnconfigure(0, weight=1)

    text_area = Text(log_frame, wrap='word', height=15, width=50, font=("Segoe UI", 10))
    text_area.grid(row=0, column=0, sticky='nsew', padx=(0, 5))

    scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=text_area.yview)
    scrollbar.grid(row=0, column=1, sticky='ns')
    text_area['yscrollcommand'] = scrollbar.set
    try:
        with open('excel_comparison.log', 'r') as log_file:
            text_area.insert('1.0', log_file.read())
            text_area.config(state='disabled')
    except IOError:
        messagebox.showerror(LANG_PL["error"], LANG_PL["could_not_open_log"])
        log_window.destroy()

def populate_columns(filepath):
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() for cell in ws[1] if cell.value]
        column_combobox['values'] = headers
        if headers:
            column_combobox.current(0)
        else:
            column_combobox.set('')
            column_combobox['values'] = []
        update_check_button()
    except Exception as e:
        messagebox.showerror(LANG_PL["error"], LANG_PL["failed_to_load_columns"].format(e=e))
        column_combobox.set('')
        column_combobox['values'] = []
        update_check_button()

def show_results(res_list):
    global last_results
    last_results = res_list
    results_tree.delete(*results_tree.get_children())
    for item in res_list:
        status_text = LANG_PL["status_found"] if item['status'] else LANG_PL["status_missing"]
        tag = 'found' if item['status'] else 'missing'
        details = item.get('found_path', '') or '; '.join(item.get('suggestions', []))
        results_tree.insert('', 'end', values=(item.get('invoice', 'N/A'), status_text, details), tags=(tag,))
    # Status label is updated via queue now, not here
    # status_label['text'] = f"Processing complete. {len(res_list)} items checked."

def show_help():
    help_win = tk.Toplevel(root)
    help_win.title(LANG_PL["help_title"])
    help_win.geometry("700x550") # Increase default size
    help_win.minsize(500, 400) # Set a minimum practical size
    help_win.resizable(True, True) # Allow resizing
 
    # Apply the defined help_font to the Text widget
    base_font_size = 16 # Increased font size again to 16pt
    base_font_family = tkfont.nametofont("TkDefaultFont")["family"]
    help_font = tkfont.Font(family=base_font_family, size=base_font_size) # Re-add definition
    bold_help_font = tkfont.Font(family=base_font_family, size=base_font_size, weight="bold")
 
    text_area = tk.Text(help_win, wrap="word", padx=10, pady=10, spacing1=5, spacing3=10, font=help_font)
    text_area.pack(expand=True, fill="both")

    # Define tags for formatting
    text_area.tag_configure("bold", font=bold_help_font)
    text_area.tag_configure("heading", font=bold_help_font, spacing1=10, spacing3=5)
    text_area.tag_configure("bullet", lmargin1="20", lmargin2="20") # Indent bullets

    # Insert text and apply tags
    help_content = LANG_PL["help_text"]
    lines = help_content.split('\n')
    for line in lines:
        stripped_line = line.strip()
        if stripped_line.startswith("**") and stripped_line.endswith("**"):
            text_area.insert(tk.END, stripped_line.strip('*') + "\n", "heading")
        elif stripped_line.startswith("*"):
            text_area.insert(tk.END, " • " + stripped_line.strip('* ').strip() + "\n", "bullet") # Use bullet char
        elif stripped_line and stripped_line[0].isdigit() and '.' in stripped_line:
             # Basic check for numbered list items
            parts = stripped_line.split(' ', 1)
            if len(parts) > 1:
                first_part = parts[0]
                rest_of_line = parts[1]
                # Make the initial part bold (e.g., "1. ", "2. ")
                bold_end_index = text_area.index(tk.END)
                text_area.insert(tk.END, first_part + " ")
                bold_start_index = text_area.index(f"{bold_end_index} linestart")
                text_area.tag_add("bold", bold_start_index, text_area.index(tk.END))
                text_area.insert(tk.END, rest_of_line + "\n", "bullet") # Indent numbered items too
            else:
                text_area.insert(tk.END, stripped_line + "\n", "bullet")
        elif stripped_line:
            text_area.insert(tk.END, stripped_line + "\n")
        else:
             text_area.insert(tk.END, "\n") # Preserve blank lines for spacing

    text_area.config(state="disabled") # Make text read-only

    close_button = ttk.Button(help_win, text="Zamknij", command=help_win.destroy)
    close_button.pack(pady=10)

    # Center the help window
    help_win.transient(root) # Keep on top of main window
    help_win.grab_set() # Modal behavior
    root.wait_window(help_win)

def update_check_button(*args):
    excel_selected = excel_file_entry.get() and os.path.isfile(excel_file_entry.get())
    dir_selected = directory_entry.get() and os.path.isdir(directory_entry.get())
    col_selected = column_combobox.get()
    ext_selected = ext_var.get()
    strategy_selected = strategy_var.get()

    if excel_selected and dir_selected and col_selected and ext_selected and strategy_selected:
        check_files_button['state'] = 'normal'
    else:
        check_files_button['state'] = 'disabled'

def export_csv():
    if not last_results:
        messagebox.showwarning(LANG_PL["export_csv_button"], LANG_PL["export_csv_no_results"])
        return

    filename = filedialog.asksaveasfilename(defaultextension='.csv', filetypes=[('CSV Files','*.csv')])
    if not filename:
        return
    try:
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([LANG_PL["results_item_header"], LANG_PL["results_status_header"], LANG_PL["results_details_header"]])
            for r in last_results:
                status_text = LANG_PL["status_found"] if r['status'] else LANG_PL["status_missing"]
                details = r.get('found_path', '') or '; '.join(r.get('suggestions', []))
                writer.writerow([r['invoice'], status_text, details])
        messagebox.showinfo(LANG_PL["export_csv_button"], LANG_PL["export_csv_success"].format(filename=os.path.basename(filename)))
    except Exception as e:
        messagebox.showerror(LANG_PL["error"], LANG_PL["export_csv_failed"].format(e=e))

root = tk.Tk()
root.title(LANG_PL["window_title"])
root.geometry("1000x800")
root.minsize(600, 500)
root.configure(bg='#e0e0e0')

root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
root.rowconfigure(1, weight=0)

style = ttk.Style()
style.theme_use('clam')

style.configure('.', font=('Segoe UI', 11), background='#e0e0e0')
style.configure('TFrame', background='#e0e0e0')
style.configure('TLabel', background='#e0e0e0', font=('Segoe UI', 11))
style.configure('TButton', font=('Segoe UI', 11, 'bold'), padding=5)
style.configure('Toolbutton', padding=5)
style.configure('TEntry', font=('Segoe UI', 11), padding=3)
style.configure('TCombobox', font=('Segoe UI', 11))
style.map('TCombobox', fieldbackground=[('readonly', 'white')])
style.configure('TCheckbutton', background='#e0e0e0', font=('Segoe UI', 11))

style.configure('Treeview', rowheight=28, font=('Segoe UI', 10), background='white', fieldbackground='white')
style.configure('Treeview.Heading', background='#4a6984', foreground='white', font=('Segoe UI', 11, 'bold'), padding=5)
style.map('Treeview.Heading', relief=[('active', 'groove'), ('!active', 'flat')])

style.configure("custom.Horizontal.TProgressbar", troughcolor='#d3d3d3', background='#4a6984')

menubar = tk.Menu(root)
filemenu = tk.Menu(menubar, tearoff=0)
filemenu.add_command(label=LANG_PL["view_log_menu"], command=display_log)
filemenu.add_separator()
filemenu.add_command(label=LANG_PL["exit_menu"], command=root.quit)
menubar.add_cascade(label=LANG_PL["file_menu"], menu=filemenu)

helpmenu = tk.Menu(menubar, tearoff=0)
helpmenu.add_command(label=LANG_PL["show_help_button"], command=show_help)
menubar.add_cascade(label=LANG_PL["help_button"], menu=helpmenu)

root.config(menu=menubar)

content_frame = ttk.Frame(root, padding="10 10 10 10")
content_frame.grid(row=0, column=0, sticky="NSEW")
content_frame.columnconfigure(0, weight=1)
content_frame.rowconfigure(0, weight=0)
content_frame.rowconfigure(1, weight=0)
content_frame.rowconfigure(2, weight=0)
content_frame.rowconfigure(3, weight=1)
content_frame.rowconfigure(4, weight=0)

input_frame = ttk.LabelFrame(content_frame, text=LANG_PL["inputs_label"], padding="10")
input_frame.grid(row=0, column=0, sticky="EW", padx=5, pady=(0, 10))
input_frame.columnconfigure(1, weight=1)

excel_file_label = ttk.Label(input_frame, text=LANG_PL["excel_file_label"])
excel_file_label.grid(row=0, column=0, sticky="W", padx=5, pady=5)
excel_file_entry = ttk.Entry(input_frame, width=60)
excel_file_entry.grid(row=0, column=1, sticky="EW", padx=5, pady=5)
excel_file_button = ttk.Button(input_frame, text=LANG_PL["select_file_button"], command=lambda: select_file(excel_file_entry))
excel_file_button.grid(row=0, column=2, sticky="E", padx=5, pady=5)
Tooltip(excel_file_entry, LANG_PL["tooltip_excel_entry"])
Tooltip(excel_file_button, LANG_PL["tooltip_excel_button"])

directory_label = ttk.Label(input_frame, text=LANG_PL["dir_label"])
directory_label.grid(row=1, column=0, sticky="W", padx=5, pady=5)
directory_entry = ttk.Entry(input_frame, width=60)
directory_entry.grid(row=1, column=1, sticky="EW", padx=5, pady=5)
directory_button = ttk.Button(input_frame, text=LANG_PL["select_dir_button"], command=lambda: select_directory(directory_entry))
directory_button.grid(row=1, column=2, sticky="E", padx=5, pady=5)
Tooltip(directory_entry, LANG_PL["tooltip_dir_entry"])
Tooltip(directory_button, LANG_PL["tooltip_dir_button"])

options_frame = ttk.LabelFrame(content_frame, text=LANG_PL["options_label"], padding="10")
options_frame.grid(row=1, column=0, sticky="EW", padx=5, pady=(0, 10))
options_frame.columnconfigure(1, weight=1)
options_frame.columnconfigure(3, weight=1)
options_frame.columnconfigure(5, weight=1)

column_label = ttk.Label(options_frame, text=LANG_PL["column_label"])
column_label.grid(row=0, column=0, sticky="W", padx=5, pady=5)
column_combobox = ttk.Combobox(options_frame, state='readonly', width=25)
column_combobox.grid(row=0, column=1, sticky="EW", padx=5, pady=5)
column_combobox.bind('<<ComboboxSelected>>', update_check_button)
Tooltip(column_combobox, LANG_PL["tooltip_column_combo"])

ext_label = ttk.Label(options_frame, text=LANG_PL["ext_label"])
ext_label.grid(row=0, column=2, sticky="W", padx=15, pady=5)
ext_var = tk.StringVar()
ext_combobox = ttk.Combobox(options_frame, textvariable=ext_var, values=['.pdf', '.txt', '.docx', '.xlsx', '.csv'], width=10)
ext_combobox.grid(row=0, column=3, sticky="EW", padx=5, pady=5)
ext_combobox.current(0)
ext_combobox.bind('<<ComboboxSelected>>', update_check_button)
Tooltip(ext_combobox, LANG_PL["tooltip_ext_combo"])

strategy_label = ttk.Label(options_frame, text=LANG_PL["strategy_label"])
strategy_label.grid(row=1, column=0, sticky="W", padx=5, pady=5)
strategy_var = tk.StringVar(value='Dokładna')
strategy_combobox = ttk.Combobox(options_frame, textvariable=strategy_var, 
                                 values=[LANG_PL["strategy_exact"], LANG_PL["strategy_contains"]], # Use Polish display names
                                 state='readonly', width=15)
strategy_combobox.grid(row=1, column=1, sticky="EW", padx=5, pady=5)
strategy_combobox.bind('<<ComboboxSelected>>', update_check_button)
Tooltip(strategy_combobox, LANG_PL["tooltip_strategy_combo"])

recurse_var = tk.BooleanVar()
recurse_check = ttk.Checkbutton(options_frame, text=LANG_PL["include_subdirs_label"], variable=recurse_var)
recurse_check.grid(row=1, column=2, columnspan=2, sticky="W", padx=15, pady=5)
Tooltip(recurse_check, LANG_PL["tooltip_subdir_check"])

action_frame = ttk.Frame(content_frame)
action_frame.grid(row=2, column=0, sticky="EW", padx=5, pady=10)
action_frame.columnconfigure(0, weight=1)

check_files_button = ttk.Button(action_frame, text=LANG_PL["check_files_button"], 
                              command=lambda: start_file_check(excel_file_entry, directory_entry, q), state='disabled')
check_files_button.grid(row=0, column=0, pady=5)
Tooltip(check_files_button, LANG_PL["tooltip_check_button"])

results_frame = ttk.LabelFrame(content_frame, text=LANG_PL["results_label"], padding="10")
results_frame.grid(row=3, column=0, sticky="NSEW", padx=5, pady=(0, 10))
results_frame.columnconfigure(0, weight=1)
results_frame.rowconfigure(0, weight=1)

results_tree = ttk.Treeview(results_frame, columns=('Item', 'Status', 'Details'), show='headings')
results_tree.heading('Item', text=LANG_PL["results_item_header"])
results_tree.heading('Status', text=LANG_PL["results_status_header"])
results_tree.heading('Details', text=LANG_PL["results_details_header"])
results_tree.column('Item', width=250, anchor='w')
results_tree.column('Status', width=100, anchor='center')
results_tree.column('Details', width=500, anchor='w')

vsb = ttk.Scrollbar(results_frame, orient="vertical", command=results_tree.yview)
hsb = ttk.Scrollbar(results_frame, orient="horizontal", command=results_tree.xview)
results_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

results_tree.grid(row=0, column=0, sticky='nsew')
vsb.grid(row=0, column=1, sticky='ns')
hsb.grid(row=1, column=0, sticky='ew')

results_tree.tag_configure('missing', background='#ffebee')
results_tree.tag_configure('found', background='#e8f5e9')

export_button = ttk.Button(content_frame, text=LANG_PL["export_csv_button"], command=export_csv)
export_button.grid(row=4, column=0, sticky="E", padx=10, pady=(0,5))
Tooltip(export_button, LANG_PL["tooltip_export_button"])

status_frame = ttk.Frame(root, padding=(5, 2, 5, 2), relief='sunken')
status_frame.grid(row=1, column=0, sticky='EW')
status_frame.columnconfigure(0, weight=1)

status_label = ttk.Label(status_frame, text=LANG_PL["status_ready"], anchor='w')
status_label.grid(row=0, column=0, sticky='EW', padx=5)

progress = ttk.Progressbar(status_frame, orient='horizontal', length=200, mode='determinate', style="custom.Horizontal.TProgressbar")
progress.grid(row=0, column=1, sticky='E', padx=5)

last_results = []

def on_result_double_click(event):
    item_id = results_tree.focus()
    if not item_id:
        return
    item = results_tree.item(item_id)
    values = item['values']
    tags = item['tags']

    if 'missing' in tags and last_results:
        missing_value = values[0]
        for result in last_results:
            if result['invoice'] == missing_value and not result['status'] and 'suggestions' in result:
                suggestions = result['suggestions']
                if suggestions:
                    messagebox.showinfo(LANG_PL["suggestions_title"], LANG_PL["suggestions_found"].format(value=missing_value, suggestions='\n- '.join(suggestions)))
                else:
                    messagebox.showinfo(LANG_PL["suggestions_title"], LANG_PL["suggestions_not_found"].format(value=missing_value))
                break

results_tree.bind("<Double-1>", on_result_double_click)

CONFIG_FILE = "config.json"

def load_config():
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r') as f:
                config = json.load(f)

            excel_path = config.get("excel_file", "")
            dir_path = config.get("directory", "")
            col_name = config.get("column", "")
            extension = config.get("extension", ".pdf")
            strategy_internal = config.get("strategy_key", "exact") # Load internal key
            subdirs = config.get("subdirs", False)

            if excel_path and os.path.isfile(excel_path):
                excel_file_entry.insert(0, excel_path)
                populate_columns(excel_path)
                if col_name in column_combobox['values']:
                    column_combobox.set(col_name)
            
            if dir_path and os.path.isdir(dir_path):
                directory_entry.insert(0, dir_path)

            if extension in ext_combobox['values']:
                ext_var.set(extension)
                
            # Map internal key back to Polish display name
            strategy_display = get_key_from_value(STRATEGY_MAP_PL, strategy_internal)
            if strategy_display and strategy_display in strategy_combobox['values']:
                 strategy_var.set(strategy_display)
            else: # Default to 'Dokładna' if loaded key is invalid or removed
                 strategy_combobox.current(0) 

            recurse_var.set(subdirs)
            update_check_button() # Update button state based on loaded config
            
    except (FileNotFoundError, json.JSONDecodeError) as e:
        # Silently ignore if config doesn't exist or is invalid on first run
        # Or show a warning: messagebox.showwarning("Config", f"Could not load config: {e}")
        pass
    except Exception as e:
        # Catch other potential errors during loading
        messagebox.showwarning(LANG_PL["config_load_error"], f"{LANG_PL['config_load_error'].format(e=e)}")

def save_config():
    strategy_display = strategy_var.get()
    strategy_internal = STRATEGY_MAP_PL.get(strategy_display, 'exact') # Get internal key for saving
    # Ensure we only save valid keys
    if strategy_internal not in ['exact', 'contains']:
        strategy_internal = 'exact'
    config = {
        "excel_file": excel_file_entry.get(),
        "directory": directory_entry.get(),
        "column": column_combobox.get(),
        "extension": ext_var.get(),
        "strategy_key": strategy_internal, # Save internal key
        "subdirs": recurse_var.get()
    }
    try:
        with open(CONFIG_FILE, 'w') as f:
            json.dump(config, f, indent=4)
    except IOError as e:
        messagebox.showerror(LANG_PL["error"], LANG_PL['config_save_error'].format(e=e))

def on_closing():
    save_config()
    root.destroy()

q = queue.Queue()
load_config() # Load config on startup
root.protocol("WM_DELETE_WINDOW", on_closing) # Save config on close
root.after(100, update_gui, q)
root.mainloop()

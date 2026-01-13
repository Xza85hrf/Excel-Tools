import os
import logging
import re
from rich.progress import track

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# --- Logging Setup ---
if os.path.exists('excel_comparison.log'):
    try:
        os.remove('excel_comparison.log')
    except OSError as e:
        print(f"Warning: Could not remove old log file: {e}")

logging.basicConfig(
    filename='excel_comparison.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# --- Fill Styles ---
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

def find_filename_column(ws, column_name="Numer faktury"):
    logging.info(f"Searching for column '{column_name}'...")
    header_row = ws[1]
    for i, cell in enumerate(header_row):
        if cell.value and str(cell.value).strip().lower() == column_name.lower():
            logging.info(f"Found '{column_name}' column at index {i}.")
            return i
    logging.error(f"'{column_name}' column not found in worksheet.")
    raise ValueError(f"Could not find '{column_name}' column in worksheet.")

def adjust_column_width(ws):
    for i, column in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        column_cells = tuple(ws.iter_cols(min_col=i, max_col=i))
        for cells_in_row in column_cells:
            cell = cells_in_row[0]
            try:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except Exception as e:
                logging.warning(f"Could not process cell {cell.coordinate}: {e}")
        adjusted_width = (max_length * 1.2) + 2
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width

# --- Utility functions for testing and CLI integration ---
def load_excel_data(excel_file_path, column_name):
    """Load list of values from given column in Excel file."""
    from openpyxl import load_workbook
    wb = load_workbook(excel_file_path)
    ws = wb.active
    col_index = find_filename_column(ws, column_name)
    items = []
    for row in ws.iter_rows(min_row=2, min_col=col_index+1, max_col=col_index+1, values_only=True):
        value = row[0]
        if value is not None:
            items.append(str(value).strip())
    return items

def run_check_files(items_to_check, directory_path, extensions, include_subdirs, match_mode, q):
    """Check presence of files matching items_to_check in directory and return structured results."""
    import os, re
    # Prepare extension filter
    exts = [e.lower() for e in extensions]
    no_filter = any(e == '' for e in exts)
    # Gather file paths
    file_paths = []
    if include_subdirs:
        for root, dirs, files in os.walk(directory_path):
            for f in files:
                if no_filter or os.path.splitext(f)[1].lower() in exts:
                    file_paths.append(os.path.join(root, f))
    else:
        for f in os.listdir(directory_path):
            p = os.path.join(directory_path, f)
            if os.path.isfile(p) and (no_filter or os.path.splitext(f)[1].lower() in exts):
                file_paths.append(p)
    results = []
    statuses = {'found': 'Znaleziono', 'missing': 'Brakujący'}
    for item in items_to_check:
        found_path = None
        for path in file_paths:
            name = os.path.splitext(os.path.basename(path))[0]
            if match_mode == 'exact' and name.lower() == item.lower():
                found_path = path; break
            elif match_mode == 'contains' and item.lower() in name.lower():
                found_path = path; break
        if found_path:
            status = statuses['found']
            details = found_path
        else:
            status = statuses['missing']
            details = ''
        results.append({'excel_value': item, 'status': status, 'details': details})
    return results

# Alias existing function to match test import
adjust_column_widths = adjust_column_width

def main(excel_file_path, directory_path, column_name, extensions, include_subdirs, match_mode, q):
    try:
        q.put(("update_label", "Loading Excel file..."))
        q.put(("update_progress", 0))
        logging.info(f"Loading workbook: {excel_file_path}")
        _, ext = os.path.splitext(excel_file_path)
        if ext.lower() not in ['.xlsx']:
            q.put(("update_label", "Error: Only .xlsx files are supported."))
            logging.error(f"Unsupported file type: {ext}")
            return False, None

        workbook = load_workbook(excel_file_path)
        ws = workbook.active
        logging.info(f"Successfully loaded workbook. Active sheet: {ws.title}")

        col_index = find_filename_column(ws, column_name)

        q.put(("update_label", "Checking files..."))
        logging.info(f"Starting file check in directory: {directory_path}")

        max_row = ws.max_row
        # Build list of files based on extensions and recursion
        exts = [e.lower() for e in extensions]
        file_paths = []
        # prepare results list for GUI summary
        results = []
        if include_subdirs:
            try:
                for root, dirs, files in os.walk(directory_path):
                    for f in files:
                        if os.path.splitext(f)[1].lower() in exts:
                            file_paths.append(os.path.join(root, f))
            except Exception as e:
                logging.warning(f"Error walking '{directory_path}': {e}")
        else:
            try:
                for f in os.listdir(directory_path):
                    if os.path.splitext(f)[1].lower() in exts:
                        file_paths.append(os.path.join(directory_path, f))
            except Exception as e:
                logging.warning(f"Error listing '{directory_path}': {e}")

        total_rows_to_process = max_row - 1 if max_row > 1 else 0
        for row_idx in track(range(2, max_row + 1), description="Processing Excel rows...", total=total_rows_to_process):
            row = ws[row_idx]
            cell = ws.cell(row=row_idx, column=col_index + 1)
            filename = str(cell.value).strip() if cell.value else None

            if not filename:
                logging.warning(f"Skipping row {row_idx}: empty filename.")
                continue

            # Select match based on strategy
            matched_path = None
            for path in file_paths:
                name = os.path.splitext(os.path.basename(path))[0]
                if match_mode == 'exact' and name.lower() == filename.lower():
                    matched_path = path; break
                elif match_mode == 'contains' and filename.lower() in name.lower():
                    matched_path = path; break
            if matched_path:
                found = True
                file_to_check = matched_path
                logging.debug(f"Row {row_idx}: {match_mode} match '{os.path.basename(matched_path)}' for '{filename}'.")
            else:
                found = False
                file_to_check = os.path.join(directory_path, filename)

            if found:
                fill = green_fill
                logging.debug(f"Row {row_idx}: File found for '{filename}'.")
            else:
                fill = red_fill
                logging.info(f"Row {row_idx}: File '{filename}' not found.")

            for c in row:
                c.fill = fill
            # record result and suggestions
            suggestions = [os.path.basename(p) for p in file_paths if filename.lower() in os.path.basename(p).lower()] if not found else []
            results.append({"invoice": filename, "status": found, "suggestions": suggestions})

            progress = int(((row_idx - 1) / total_rows_to_process) * 100) if total_rows_to_process > 0 else 100
            q.put(("update_progress", progress))

        # send summary results to GUI
        q.put(("results", results))
        logging.info("File check complete.")
        
        # -- Add 'Results' sheet for reporting --
        from datetime import datetime
        res_sheet = workbook.create_sheet(title="Results")
        # Write summary stats
        total = len(results)
        found_count = sum(1 for r in results if r["status"])
        missing_count = total - found_count
        res_sheet.append(["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        res_sheet.append(["Total", total])
        res_sheet.append(["Found", found_count])
        res_sheet.append(["Missing", missing_count])
        res_sheet.append([])
        res_sheet.append(["Invoice", "Status", "Suggestions"])
        for r in results:
            status_text = "Found" if r["status"] else "Missing"
            sugg_text = "; ".join(r.get("suggestions", []))
            res_sheet.append([r["invoice"], status_text, sugg_text])

        q.put(("update_progress", 100))
        q.put(("update_status", "Processing complete. Ready to save."))
        return True, workbook

    except ValueError as ve:
        logging.error(f"Config error: {ve}")
        q.put(("update_label", f"Error: {ve}"))
        return False, None
    except Exception as e:
        logging.exception(f"Unexpected error: {e}")
        q.put(("update_label", f"Error: {e}"))
        return False, None

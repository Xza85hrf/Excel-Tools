import unittest
import os
import sys
import shutil
from openpyxl import Workbook
import queue

# Add the parent directory (project root) to the Python path
# This allows importing excel_comparison module
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, project_root)

# Now import the functions from the main script
# Note: We might need to refactor excel_comparison.py slightly if it relies
# heavily on global GUI elements or assumes it's run as __main__.
# For now, let's assume we can import necessary functions.
from excel_comparison import load_excel_data, run_check_files, adjust_column_widths

class TestExcelComparison(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        """Set up test environment once for the class."""
        cls.test_dir = os.path.join(project_root, 'tests')
        cls.excel_file_path = os.path.join(cls.test_dir, 'test_list.xlsx')
        cls.files_dir = os.path.join(cls.test_dir, 'test_files_to_check')

        # Create dummy Excel file for testing
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws['A1'] = "Irrelevant Header"
        ws['B1'] = "Item Name" # The column we'll use
        ws['C1'] = "Another Header"
        # --- Data ---
        ws['B2'] = "Item1"
        ws['B3'] = "Item2"
        ws['B4'] = "Item3"
        ws['B5'] = "Item4"
        ws['B6'] = "MissingItem" # This item doesn't have a corresponding file
        ws['B7'] = "ITEM2" # Test case sensitivity if needed later

        # Ensure the directory exists before saving
        os.makedirs(os.path.dirname(cls.excel_file_path), exist_ok=True)
        wb.save(cls.excel_file_path)

        # Dummy file structure should already be created by previous steps
        # We just verify the base directory exists
        if not os.path.exists(cls.files_dir):
             # Recreate if missing (shouldn't happen based on previous steps)
            os.makedirs(os.path.join(cls.files_dir, 'subfolder'), exist_ok=True)
            with open(os.path.join(cls.files_dir, 'Item1.pdf'), 'w') as f: f.write('dummy')
            with open(os.path.join(cls.files_dir, 'ITEM2.TXT'), 'w') as f: f.write('dummy')
            with open(os.path.join(cls.files_dir, 'PrefixItem3Suffix.pdf'), 'w') as f: f.write('dummy')
            with open(os.path.join(cls.files_dir, 'subfolder', 'Item4.pdf'), 'w') as f: f.write('dummy')
            with open(os.path.join(cls.files_dir, 'WrongItem.txt'), 'w') as f: f.write('dummy')


    @classmethod
    def tearDownClass(cls):
        """Clean up test environment once after all tests are run."""
        # Remove the generated Excel file
        if os.path.exists(cls.excel_file_path):
            os.remove(cls.excel_file_path)
        # Optionally remove dummy files/dirs if needed, but often kept for inspection
        # if os.path.exists(cls.files_dir):
        #     shutil.rmtree(cls.files_dir)

    def test_01_load_excel_data(self):
        """Test loading data from the specified column in the Excel file."""
        items_to_check = load_excel_data(self.excel_file_path, "Item Name")
        expected_items = ["Item1", "Item2", "Item3", "Item4", "MissingItem", "ITEM2"]
        self.assertListEqual(sorted(items_to_check), sorted(expected_items))

    def test_02_run_check_files_exact_no_subdir_pdf(self):
        """Test exact match, PDF extension, no subdirectories."""
        items_to_check = ["Item1", "Item2", "Item3", "Item4", "MissingItem"]
        results = run_check_files(items_to_check, self.files_dir, [".pdf"], False, "exact", queue.Queue())
        found = {item['excel_value'] for item in results if item['status'] == 'Znaleziono'}
        missing = {item['excel_value'] for item in results if item['status'] == 'Brakujący'}
        self.assertEqual(found, {"Item1"})
        self.assertEqual(missing, {"Item2", "Item3", "Item4", "MissingItem"})
        # Check details for found item
        item1_details = next(item['details'] for item in results if item['excel_value'] == 'Item1')
        self.assertTrue(item1_details.endswith('Item1.pdf'))

    def test_03_run_check_files_exact_with_subdir_pdf(self):
        """Test exact match, PDF extension, including subdirectories."""
        items_to_check = ["Item1", "Item2", "Item3", "Item4", "MissingItem"]
        results = run_check_files(items_to_check, self.files_dir, [".pdf"], True, "exact", queue.Queue())
        found = {item['excel_value'] for item in results if item['status'] == 'Znaleziono'}
        missing = {item['excel_value'] for item in results if item['status'] == 'Brakujący'}
        self.assertEqual(found, {"Item1", "Item4"})
        self.assertEqual(missing, {"Item2", "Item3", "MissingItem"})
        # Check details for found items
        item1_details = next(item['details'] for item in results if item['excel_value'] == 'Item1')
        item4_details = next(item['details'] for item in results if item['excel_value'] == 'Item4')
        self.assertTrue(item1_details.endswith('Item1.pdf'))
        self.assertTrue(item4_details.endswith(os.path.join('subfolder', 'Item4.pdf')))

    def test_04_run_check_files_contains_no_subdir_pdf(self):
        """Test contains match, PDF extension, no subdirectories."""
        items_to_check = ["Item1", "Item2", "Item3", "Item4", "MissingItem"]
        results = run_check_files(items_to_check, self.files_dir, [".pdf"], False, "contains", queue.Queue())
        found = {item['excel_value'] for item in results if item['status'] == 'Znaleziono'}
        missing = {item['excel_value'] for item in results if item['status'] == 'Brakujący'}
        # Expect Item1 (Item1.pdf) and Item3 (PrefixItem3Suffix.pdf)
        self.assertEqual(found, {"Item1", "Item3"})
        self.assertEqual(missing, {"Item2", "Item4", "MissingItem"})
        item1_details = next(item['details'] for item in results if item['excel_value'] == 'Item1')
        item3_details = next(item['details'] for item in results if item['excel_value'] == 'Item3')
        self.assertTrue(item1_details.endswith('Item1.pdf'))
        self.assertTrue(item3_details.endswith('PrefixItem3Suffix.pdf'))

    def test_05_run_check_files_contains_with_subdir_pdf(self):
        """Test contains match, PDF extension, including subdirectories."""
        items_to_check = ["Item1", "Item2", "Item3", "Item4", "MissingItem"]
        results = run_check_files(items_to_check, self.files_dir, [".pdf"], True, "contains", queue.Queue())
        found = {item['excel_value'] for item in results if item['status'] == 'Znaleziono'}
        missing = {item['excel_value'] for item in results if item['status'] == 'Brakujący'}
        # Expect Item1, Item3, Item4
        self.assertEqual(found, {"Item1", "Item3", "Item4"})
        self.assertEqual(missing, {"Item2", "MissingItem"})

    def test_06_run_check_files_exact_no_subdir_txt(self):
        """Test exact match, TXT extension, no subdirectories."""
        items_to_check = ["Item1", "Item2", "MissingItem"]
        # Note: File is ITEM2.TXT, search is for Item2
        results = run_check_files(items_to_check, self.files_dir, [".txt"], False, "exact", queue.Queue())
        found = {item['excel_value'] for item in results if item['status'] == 'Znaleziono'}
        missing = {item['excel_value'] for item in results if item['status'] == 'Brakujący'}
        self.assertEqual(found, {"Item2"})
        self.assertEqual(missing, {"Item1", "MissingItem"})
        item2_details = next(item['details'] for item in results if item['excel_value'] == 'Item2')
        # Exact match is case-insensitive on filename comparison in the function
        self.assertTrue(item2_details.endswith('ITEM2.TXT'))

    def test_07_run_check_files_no_extension_filter(self):
        """Test exact match, no extension filter, including subdirectories."""
        items_to_check = ["Item1", "Item2", "Item3", "Item4", "MissingItem"]
        results = run_check_files(items_to_check, self.files_dir, [""], True, "exact", queue.Queue())
        found = {item['excel_value'] for item in results if item['status'] == 'Znaleziono'}
        missing = {item['excel_value'] for item in results if item['status'] == 'Brakujący'}
        # Expect Item1.pdf, ITEM2.TXT, Item4.pdf (Item3 not found because exact match fails)
        self.assertEqual(found, {"Item1", "Item2", "Item4"})
        self.assertEqual(missing, {"Item3", "MissingItem"})

    def test_08_run_check_files_result_structure(self):
        """Verify the structure of the results list."""
        items_to_check = ["Item1"]
        results = run_check_files(items_to_check, self.files_dir, [".pdf"], False, "exact", queue.Queue())
        self.assertEqual(len(results), 1)
        result = results[0]
        self.assertIn('excel_value', result)
        self.assertIn('status', result)
        self.assertIn('details', result)
        self.assertEqual(result['excel_value'], 'Item1')
        self.assertEqual(result['status'], 'Znaleziono')
        self.assertTrue(result['details'].endswith('Item1.pdf'))

    # Note: Testing adjust_column_widths might require creating a dummy workbook 
    # in memory and checking ColumnDimension objects, or saving and reloading.
    # This adds complexity and dependency on file I/O within the test.
    # For now, we'll skip direct testing of adjust_column_widths but ensure
    # the main run_check_files logic works.

if __name__ == '__main__':
    unittest.main()

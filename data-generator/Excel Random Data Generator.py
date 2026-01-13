from tkinter import Tk, Label, Entry, Button

import faker
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

# Create Faker object
fake = faker.Faker()


def generate_random_data(num_records, headers=None):
    if headers is None:
        headers = ["Client nr", "Client", "Address", "NIP"]

    data = {headers[0]: [j for j in range(1, num_records + 1)], headers[1]: [fake.name() for _ in range(num_records)],
            headers[2]: [fake.address().replace('\n', ', ') for _ in range(num_records)],
            headers[3]: [fake.random_int(min=1000000000, max=9999999999) for _ in range(num_records)]}
    return pd.DataFrame(data)


def adjust_column_width(ws):
    for column_index, column_cells in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        column = [cell for cell in column_cells]
        for cell in column:
            try:
                cell_length = len(str(cell.value))
            except ValueError:
                pass
            else:
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = (max_length * 1.2)
        ws.column_dimensions[get_column_letter(column_index)].width = adjusted_width


def save_data_to_excel(data, filename):
    try:
        data.to_excel(filename, index=False)
        wb = load_workbook(filename)
        ws = wb.active
        adjust_column_width(ws)
        wb.save(filename)
    except (FileNotFoundError, PermissionError) as e:
        print(f"Error while writing to {filename}: {e}")


def generate_and_save():
    headers = headers_entry.get().split(',')
    for index, entry in enumerate(num_clients_entries):
        num_clients = int(entry.get())
        if num_clients > 0:
            data = generate_random_data(num_clients, headers)
            save_data_to_excel(data, f"excel_{num_clients}_clients_{index + 1}.xlsx")


def generate_default():
    headers = ["Client nr", "Client", "Address", "NIP"]
    data_20_clients = generate_random_data(20, headers)
    data_30_clients = generate_random_data(30, headers)
    save_data_to_excel(data_20_clients, "excel_20_clients_default.xlsx")
    save_data_to_excel(data_30_clients, "excel_30_clients_default.xlsx")


root = Tk()
root.title("Excel Random Data Generator")

Label(root, text="Enter the headers (comma-separated)").grid(row=0, column=0, columnspan=4)
headers_entry = Entry(root, width=50)
headers_entry.grid(row=1, column=0, columnspan=4)

Label(root, text="Enter the number of clients for up to 4 sheets").grid(row=2, column=0, columnspan=4)
num_clients_entries = [Entry(root, width=10) for _ in range(4)]
for i, num_clients_entry in enumerate(num_clients_entries):
    num_clients_entry.grid(row=3, column=i)

Button(root, text="Generate and save", command=generate_and_save).grid(row=4, column=0, columnspan=2)
Button(root, text="Generate Default", command=generate_default).grid(row=4, column=2, columnspan=2)

Label(root, text="Instructions: Enter headers and number of clients, then click 'Generate and save'").grid(row=5,
                                                                                                           column=0,
                                                                                                           columnspan=4)

root.mainloop()
